# Change formulas in specific cells in a cashflow spreadsheet to roll over to a new month.
import os
from openpyxl import Workbook, load_workbook
#
# specify current path
cDir = '/onedri~1/progra~1/python3/pychar~1/cashfl~1/'
cPath = os.getenv('USERPROFILE') + cDir
os.chdir(cPath)
#
# specify Excel Workbook filename
fileName = 'current_Chipman_Family_Spending_Plan.xlsm'
#
# concatenate the current path and the workbook filename; keep any VBA macros
try:
    workBook = load_workbook(cPath + '/' + fileName,keep_vba=True)
except PermissionError:
    exit('WorkBook already open in Excel - close the file & try again.')
#
# extract all worksheet names from the workbook
allWorksheets = workBook.sheetnames
#
# specify the worksheet names & properties
oldMonth = 'October'
newMonth = 'November'
oldSheetName = oldMonth + ' Cash Flow'
if oldSheetName not in allWorksheets:
    exit('***** ERROR ***** Worksheet: [' + oldSheetName + '] does not exist...')
newSheetName = newMonth + ' Cash Flow'
#
# copy sheet & set new sheet properties
oldSheet = workBook[oldSheetName]
newSheet = workBook.copy_worksheet(oldSheet)
newSheet.title = newSheetName
newSheet.sheet_properties.tabColor = 'FFFF00'
#
cellList = ['E4','E12','E13','E15','E16','E17','E18','E19','E20','E21','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38','E39','E53']
for cellCoord in cellList:
    oldFormula = newSheet[cellCoord].value
    newFormula = oldFormula.replace(oldMonth, newMonth)
    newSheet[cellCoord].value = newFormula
workBook.save(fileName)
#
#print(sheet['E12'])
# =SUMIF('October Transaction Register'!$D$5:$D$99,'Expense Category List'!$A$5,'October Transaction Register'!$G$5:$G$99)
# >>> type(sheet['E12'].value)
# <class 'str'>
#
# >>> sheet['E10'].value
# >>> type(sheet['E10'].value)
# <class 'NoneType'>
#
# Report:
# Uncomment to debug
#print('Active file: ')
#print(cPath + '/' + fileName)
#print('\n')
#print('List of all worksheets: ')
#print(allWorksheets)
#print('\n')
#print('Current (active) worksheet: ')
#print(newSheet)
